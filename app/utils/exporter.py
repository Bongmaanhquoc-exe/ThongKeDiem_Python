# ==============================================================
# exporter.py — Xuất dữ liệu ra file Excel
# ==============================================================
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from tkinter import filedialog


def xuat_excel(du_lieu, ten_file='ket_qua.xlsx', ten_cot=None):
    """
    Xuất list dict ra file Excel.
    ten_cot: dict ánh xạ key → tên cột hiển thị, ví dụ {'full_name': 'Họ tên'}
             Nếu None thì dùng key gốc làm tiêu đề.
    Trả về đường dẫn file đã lưu, hoặc None nếu người dùng huỷ.
    """
    duong_dan = filedialog.asksaveasfilename(
        defaultextension='.xlsx',
        filetypes=[('Excel files', '*.xlsx'), ('All files', '*.*')],
        initialfile=ten_file,
        title='Chọn nơi lưu file Excel',
    )
    if not duong_dan:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dữ liệu"

    if not du_lieu:
        wb.save(duong_dan)
        return duong_dan

    cac_key = list(du_lieu[0].keys())

    # Hàng tiêu đề — nền xanh đậm, chữ trắng
    for col, key in enumerate(cac_key, 1):
        tieu_de = (ten_cot or {}).get(key, key)
        cell = ws.cell(row=1, column=col, value=tieu_de)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2C3E50')
        cell.alignment = Alignment(horizontal='center')

    # Dữ liệu
    for row_idx, dong in enumerate(du_lieu, 2):
        for col_idx, key in enumerate(cac_key, 1):
            ws.cell(row=row_idx, column=col_idx, value=dong.get(key, ''))

    # Tự chỉnh độ rộng cột
    for col in ws.columns:
        do_rong = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(do_rong + 4, 40)

    wb.save(duong_dan)
    return duong_dan
